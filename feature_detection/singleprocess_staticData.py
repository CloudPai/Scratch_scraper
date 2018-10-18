# -*- coding: utf-8 -*-
import sys
from antlr4 import *
# from AntlrLexer import AntlrLexer
# from AntlrParser import AntlrParser
# from AntlrListener import AntlrListener
# from .ANTLR import AntlrLexer
# from .ANTLR import AntlrParser
# from .ANTLR import AntlrListener
from feature_detection.AntlrLexer import AntlrLexer
from feature_detection.AntlrListener import AntlrListener
from feature_detection.AntlrParser import AntlrParser
import zipfile
import os
import codecs
import multiprocessing
import time
import xlwt
# workbook相关
from openpyxl.workbook import Workbook
# ExcelWriter，封装了很强大的excel写的功能
from openpyxl.writer.excel import ExcelWriter
# 一个eggache的数字转为列字母的方法
from openpyxl.utils import get_column_letter
from openpyxl.reader.excel import load_workbook
# import operator
import time
# from numba import jit
from timeit import timeit
import json
from collections import OrderedDict

class HandleExcel():
    def __init__(self,foldername = 'foldername',username = 'username'):
        self.head_row_labels = ['Name', '角色数(sprits_num)', '脚本数(scripts_num)', '1发送器->n接收器(shotgun_surgery_count)',
                                'n发送器->1接收器(divergent_change_count)','重复代码数(duplicate_code_count)','自定义函数的个数(differentInterfaces_num)','重复字符串数(duplicate_string_count)','全局变量数(broad_variabl_scope_num)',
                                '无意义命名数量(uncommunicativeName_count)','死代码块数(unreachableCode_count)','未初始化编程块数(blockNotInit_num)','存在于两个角色中的全局变量数(featureEnvy_num)','仅存在于一个角色的全局变量个数(temporaryField_num)',
                                '三个及以上的角色使用相同全局变量的个数(messageChains_num)',
                                '过长的脚本数(longScript_num)','无操作的脚本数(No_op_num)','未被定义的块个数(undefinedBlock_num)','未被使用的变量数(unusedVariable_num)','调用变量过多的脚本数(inappropriateIntimacy_num)',
                                '接口类型脚本数(middleMan_num)','使用的字符串数(dataClumps_num)','数据、事件、控制、运算块总数(lazyClass_num)','注释个数(comments_num)','blocks_num',

                                '1发送器->n接收器(shotgun_surgery_count)发送器list','n发送器->1接收器(divergent_change_count)接收器list','重复代码数(duplicate_code_count)脚本list',
                                '自定义函数的个数(differentInterfaces_num)list','重复字符串数(duplicate_string_count)list','全局变量数(broad_variabl_scope_num)list',
                                '无意义命名数量(uncommunicativeName_count)list','死代码块数(unreachableCode_count)list','未初始化编程块数(blockNotInit_num)list',
                                '存在于两个角色中的全局变量数(featureEnvy_num)list', '仅存在于一个角色的全局变量个数(temporaryField_num)list',
                                '三个及以上的角色使用相同全局变量的个数(messageChains_num)list','过长的脚本数(longScript_num)list','无操作的脚本数(No_op_num)list','未被定义的块个数(undefinedBlock_num)list',
                                '未被使用的变量数(unusedVariable_num)list','调用变量过多的脚本数(inappropriateIntimacy_num)list','接口类型脚本数(middleMan_num)list','使用的字符串数(dataClumps_num)list',
                                '数据、事件、控制、运算块总数(lazyClass_num)list','注释个数(comments_num)list','异常指标list(sorted_result_list)',
                                '需优先修正的指标'


                                ]
        self.staticData_threshold = {
            'Name':"threshold",
            '角色数(sprits_num)': 11,
            '脚本数(scripts_num)': 38.4,
            '1发送器->n接收器(shotgun_surgery_count)': 3,
            'n发送器->1接收器(divergent_change_count)': 1,
            '重复代码数(duplicate_code_count)': 0.2,
            '自定义函数的个数(differentInterfaces_num)': 0,
            '重复字符串数(duplicate_string_count)': 2.2,
            '全局变量数(broad_variabl_scope_num)': 1,
            '无意义命名数量(uncommunicativeName_count)': 3,
            '死代码块数(unreachableCode_count)': 2,
            '未初始化编程块数(blockNotInit_num)': 3,
            '存在于两个角色中的全局变量数(featureEnvy_num)': 0,
            '仅存在于一个角色的全局变量个数(temporaryField_num)': 0,
            '三个及以上的角色使用相同全局变量的个数(messageChains_num)': 0,
            '过长的脚本数(longScript_num)': 4,
            '无操作的脚本数(No_op_num)': 0,
            '未被定义的块个数(undefinedBlock_num)': 0,
            '未被使用的变量数(unusedVariable_num)': 1,
            '调用变量过多的脚本数(inappropriateIntimacy_num)': 0,
            '接口类型脚本数(middleMan_num)': 5,
            '使用的字符串数(dataClumps_num)': 9.4,
            '数据、事件、控制、运算块总数(lazyClass_num)': 124.6,
            '注释个数(comments_num)': 0
        }

        self.path = os.path.abspath('.')
        # self.filepath = self.path + '/test2'
        self.filepath = './sb2Files/'+str(foldername)

        # self.save_name = "Scratch_test2.xlsx"
        self.save_name = './excelFiles/'+str(username)+ ".xlsx"



    def read_from_file(self):
        all_count = 0
        blockCount = {}
        pathDir = os.listdir(self.filepath)
        for allDir in pathDir:

            try:

                child = os.path.join(self.filepath, allDir)

                print("name=", allDir)
                listener_blockCount = gen(child)
                blockCount[allDir] = listener_blockCount

                all_count += 1
                print("No:" + str(all_count))

            except:
                pass
            continue
        print('成功读取sb2文件数量：' + str(all_count) + "个！")
        return blockCount

    def write_to_excel_with_openpyxl(self, records, head_row, save_excel_name):
        # 新建一个workbook
        wb = Workbook()
        # 新建一个excelWriter
        ew = ExcelWriter(workbook=wb)
        # 设置文件输出路径与名称
        dest_filename = save_excel_name
        # 第一个sheet是ws
        ws = wb.worksheets[0]
        # 设置ws的名称
        ws.title = "range names"
        # 写第一行，标题行
        for h_x in range(1, len(head_row) + 1):
            h_col = get_column_letter(h_x)
            ws.cell('%s%s' % (h_col, 1)).value = '%s' % (head_row[h_x - 1])
            # 写第二行的阈值
            ws.cell('%s%s' % (h_col, 2)).value = '%s' % (self.staticData_threshold.get(head_row[h_x - 1]))






        # 写第三行及其以后的那些行
        row = 2
        for name in records:
            ws.cell('%s%s' % ('A', row)).value = '%s' % name.replace("_","/").replace(".sb2","")
            col = 2
            err_list = []
            err_dict = {}
            output_list = []
            needfixed_dict = {}
            output_dict = {}
            for point in records[name]:


                col_num = get_column_letter(col)
                ws.cell('%s%s' % (col_num, row)).value = '%s' % records[name][point]

                if(col <=24):
                    output_list.append(int(records[name][point]))
                    output_dict[str(ws.cell(row=1,column=col).value)] = int(records[name][point])
                    num = float(records[name][point])
                    threshold = float(ws.cell(row=2,column=col).value)
                    if(col > 3 and num > threshold ):
                        err = (num - threshold)
                        if(err > 0.5 * threshold):
                            # err_list.append(head_row[col])
                            err_list.append(ws.cell(row=1,column=col).value+"误差率:"+str(err))
                            err_dict[ws.cell(row=1,column=col).value] = err




                col += 1
            err_dict = sorted(err_dict.items(),key=lambda d: d[1],reverse=True)
            ws.cell('%s%s' % (get_column_letter(col), row)).value = '%s' % str(err_dict)
            col += 1
            try:
                needfixed = str(err_dict[0]).split(',')[0][1:].replace("'","")
                needfixed_dict[needfixed] = float(str(err_dict[0]).split(',')[1][:-1])
            except:
                pass

            ws.cell('%s%s' % (get_column_letter(col), row)).value = '%s' % str(needfixed_dict)
            col += 1

            # ws.cell('%s%s' % (get_column_letter(col), row)).value = '%s' % str(output_dict)
            row += 1

        ew.save(filename=dest_filename)
        # wb.save(filename=dest_filename)

    def run_main_save_to_excel_with_openpyxl(self):
        dataset_list = self.read_from_file()
        '''test use openpyxl to handle EXCEL 2007'''
        head_row_label = self.head_row_labels
        self.write_to_excel_with_openpyxl(dataset_list, head_row_label, self.save_name)


def unzip_scratch(filename):
    """
    unzip scratch project and extract project.json file
    :param filename: filename fo scratch project
    :return: null or project.json content
    """
    zfile = zipfile.ZipFile(filename, 'r')
    if "project.json" in zfile.namelist():
        data = zfile.read("project.json")
        return data
    else:
        return None


def gen(argv):
    raw_json = unzip_scratch(argv)
    encoded_json = codecs.decode(raw_json, 'utf-8', 'strict')

    input = InputStream(encoded_json)
    if not input:
        return

    lexer = AntlrLexer(input)
    stream = CommonTokenStream(lexer)
    parser = AntlrParser(stream)
    tree = parser.json()
    walker = ParseTreeWalker()
    listener = AntlrListener()
    walker.walk(listener, tree)
    # print("listener:", listener.blockCount, listener.dynamicDataCount)
    # print(listener.blockLocation)

    return listener.blockCount


# if __name__ == '__main__':
#     begin_time = time.time()
#     p = HandleExcel(username='username')
#     p.run_main_save_to_excel_with_openpyxl()
#
#     end_time = time.time()
#
#     total_sec = end_time - begin_time
#
#     print("总共耗时(s): " + str(total_sec))



